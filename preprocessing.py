import numpy as np
import pandas as pd
import ast
import json
import openpyxl
from sklearn.preprocessing import StandardScaler
from var_def import (iron, neuro_all, neuro_mean, neuro_median, neuro_combined, behavior, psychopathology)

# read xlsx file
OSPAN = pd.read_csv('./Data/Original_Data/OSPAN_data_5.7.24.csv', header=1)
data = pd.read_excel('./Data/Original_Data/CC_request_DrWorthy_05302024.xlsx')
# workbook = openpyxl.load_workbook('./Data/Original_Data/CC_request_DrWorthy_DICTIONARY_05302024.xlsx')
workbook = openpyxl.load_workbook('./Data/Original_Data/CC_request_DrWorthy_DICTIONARY_05302024 - Copy.xlsx')
sheet = workbook.active
questionnaires = pd.read_excel('./Data/Original_Data/A&M_dataset.xlsx')
questionnaires_dict = pd.read_excel('./Data/Original_Data/A&M_dataset_DICTIONARY.xlsx', sheet_name='Dataset_Dictionary')
questionnaires_totals = pd.read_excel('./Data/Original_Data/A&M_dataset_DICTIONARY.xlsx', sheet_name='Scores')

SGT = []

with open('./Data/Original_Data/jatos_results_Pulled_4_10_24.txt', 'r') as file:
    for line in file:
        json_data = json.loads(line)
        SGT.append(json_data)

SGT = pd.DataFrame(SGT)

choice = "else"

# =============================================================================
#                               Data Cleaning
# =============================================================================
# main data
# filter out excluded participants
data = data[data['Group'] != "Excluded"]

# also exlude the participants with high hsCRP
data = data[data['High_hsCRP'] != 1]

# remove columns with too many missing values
missing = data.isna().sum().sort_values(ascending=False)
missing_cols = missing[missing > 0.9 * data.shape[0]].index

# Ensure that the columns to be dropped exist in the DataFrame
drop_cols = list(set(missing_cols) & set(data.columns)) + ['Visit']
drop_cols = [col for col in drop_cols if col in data.columns]

data = data.drop(columns=drop_cols)

# detect the missing values and replace all -999 with NaN
data = data.replace(-999.0, np.nan)
data = data.replace(-9999.0, np.nan)

# # drop rows with missing values
data = data.dropna()

# =============================================================================
# the following code is to select variables of interest
# =============================================================================
# average the left and right hemisphere volumes
data['Putamen'] = (data['r_Pu_mean'] + data['l_Pu_mean']) / 2
data['Caudate'] = (data['r_Cd_mean'] + data['l_Cd_mean']) / 2
data['Globus_Pallidus'] = (data['r_GP_mean'] + data['l_GP_mean']) / 2

# calculate the time difference between the two recall trials
data['RAVLT_ListA_Delay_Recall_Time1'] = pd.to_datetime(data['RAVLT_ListA_Delay_Recall_Time1'])
data['RAVLT_ListA_Delay_Recall_Time2'] = pd.to_datetime(data['RAVLT_ListA_Delay_Recall_Time2'])
data['RAVLT_ListA_RecallDelay_Trial2'] = (data['RAVLT_ListA_Delay_Recall_Time2'] -
                                          data['RAVLT_ListA_Delay_Recall_Time1']).dt.total_seconds() / 60

# remove the two columns
data = data.drop(columns=['RAVLT_ListA_Delay_Recall_Time1', 'RAVLT_ListA_Delay_Recall_Time2'])

# update the dropped columns
drop_cols.extend(['RAVLT_ListA_Delay_Recall_Time1', 'RAVLT_ListA_Delay_Recall_Time2'])


# Function to check if a cell is highlighted
def is_highlighted(cell):
    if cell.fill.start_color.index == '00000000':
        return False
    return True


# Iterate through the cells and find highlighted ones
highlighted_cells = []
reverse_cols = []

for row in sheet.iter_rows():
    for cell in row:
        if is_highlighted(cell):
            highlighted_cells.append(cell)

# Iterate through the cells and find the reverse-coded ones
for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip headers
    for cell in row:
        if cell.value == 'r':
            first_column_value = row[0].value  # Accessing the first column value in the same row
            if first_column_value is not None:
                reverse_cols.append(first_column_value)

# store the highlighted variables as variables of interest
variables_of_interest = [cell.value for cell in highlighted_cells]
additional_variables = ['Putamen', 'Caudate', 'Globus_Pallidus']
variables_of_interest.extend(additional_variables)

# if there are variables of interest that do not exist in the data, print and store them
missing_vars = [var for var in variables_of_interest if var not in data.columns and var not in drop_cols]

# select the variables of interest
variables_of_interest = [var for var in variables_of_interest if var in data.columns]
data = data[variables_of_interest]

# reverse code the variables by subtracting them from the maximum value
for col in reverse_cols:
    if col in data.columns:
        data[col] = data[col].max() - data[col]
    else:
        print(f'{col} not in data')

# checkers - print the max and min values of the columns
# for col in data.columns:
#     print(col, data[col].max(), data[col].min())

# print the missing variables
print(f'Missing variables: {missing_vars}')

# =============================================================================
# the following code is to prepare data for MATLAB PLS implementation
# =============================================================================
# separate by group
psych_pls = data[data['Group'] == 'Psychiatric']
control_pls = data[data['Group'] == 'Control']

# remove the ID and Group columns
psych_pls = psych_pls.drop(columns=['Code', 'Group'])
control_pls = control_pls.drop(columns=['Code', 'Group'])

# separate the neuroimaging data
psych_without_neuro = psych_pls.drop(columns=neuro_all)
control_without_neuro = control_pls.drop(columns=neuro_all)

psych_behav = psych_pls[behavior]
control_behav = control_pls[behavior]

psych_mean = psych_pls[neuro_mean]
control_mean = control_pls[neuro_mean]

psych_median = psych_pls[neuro_median]
control_median = control_pls[neuro_median]

psych_combined = psych_pls[neuro_combined]
control_combined = control_pls[neuro_combined]

psych_with_mean = pd.concat([psych_without_neuro, psych_mean], axis=1)
control_with_mean = pd.concat([control_without_neuro, control_mean], axis=1)

psych_with_median = pd.concat([psych_without_neuro, psych_median], axis=1)
control_with_median = pd.concat([control_without_neuro, control_median], axis=1)

psych_with_combined = pd.concat([psych_without_neuro, psych_combined], axis=1)
control_with_combined = pd.concat([control_without_neuro, control_combined], axis=1)

# -----------------------------------------------------------------------------
# prepare for behavioral PLS
# -----------------------------------------------------------------------------
# separate behavioral data
psychopathology = data[psychopathology]
behavioral = data[behavior]
neuro_mean_df = data[neuro_mean]
neuro_combined_df = data[neuro_combined]
iron = data[iron]

# check if the column values are all the same for the behavioral data
uniform_vars = []
for col in behavioral.columns:
    if len(behavioral[col].unique()) == 1:
        # drop the column and print the column name
        behavioral = behavioral.drop(columns=[col])
        uniform_vars.append(col)
        print(f'{col} dropped')

print(f'Uniform variables: {uniform_vars}')
print('=' * 50)

# =============================================================================
# With additional questionnaires data, we prepare for the bifactor model
# =============================================================================

print('Preparing for bifactor model...')


# rename the columns
def rename_columns(columns, cbcl_cols=True):
    renamed_columns = []
    for i in range(len(columns)):
        if cbcl_cols:
            if i < 55:
                renamed_columns.append(f'Q{i + 1}')
            elif i == 55:
                renamed_columns.extend([f'Q56{chr(j)}' for j in range(97, 105)])  # Q56a to Q56h
            else:
                renamed_columns.append(f'Q{i + 1}') if i + 1 < (len(columns) - 6) else None
        else:
            renamed_columns.append(f'Q{i + 1}')
    return renamed_columns


def high_corr_checker(df, cutoff=0.85, checker=False):
    # check the correlation between the questionnaire items
    corr = df.corr()

    # find the variables that are highly correlated
    high_corr = []
    for i in range(corr.shape[0]):
        for j in range(i + 1, corr.shape[0]):
            if abs(corr.iloc[i, j]) > cutoff:
                high_corr.append((corr.index[i], corr.columns[j]))

    if checker:
        if high_corr:
            print('Highly correlated variables after processing:')
            for var in high_corr_checker:
                print(var)
        else:
            print('No more highly correlated variables after dropping')

    else:
        # print the highly correlated variables
        if high_corr:
            print('Highly correlated variables:')
            for var in high_corr:
                print(var)

        # take the mean of the highly correlated variables and drop the original variables
        for var in high_corr:
            df[var[0]] = ((df[var[0]] + df[var[1]]) / 2).round()
            df = df.drop(columns=[var[1]])

    return df

# only select participants that are in the data
questionnaires = questionnaires[questionnaires['Code'].isin(data['Code'])]

if choice == "CBCL":
    # select only the CBCL columns
    cbcl_cols = questionnaires.columns[questionnaires.columns.str.contains('Q')]
    questionnaires = questionnaires[cbcl_cols]

    questionnaires.columns = rename_columns(cbcl_cols)

    # remove designated columns
    questionnaires = questionnaires.drop(columns=['Q2', 'Q59', 'Q67', 'Q73', 'Q96', 'Q99', 'Q101', 'Q105'])

    # reconstruct designated columns
    def reconstruct(col1, col2, new_col, df, round=True):
        df[new_col] = (df[col1] + df[col2]) / 2
        if round:
            df[new_col] = df[new_col].round()
        else:
            df[new_col] = df[new_col]
        df = df.drop(columns=[col1, col2])
        return df


    questionnaires = reconstruct('Q20', 'Q21', 'Destroy', questionnaires, round=False)
    questionnaires = reconstruct('Q8', 'Q78', 'Inattentive', questionnaires)
    questionnaires = reconstruct('Q53', 'Q55', 'Overweight', questionnaires)

    uniform_vars = []
    for col in questionnaires.columns:
        if len(questionnaires[col].unique()) == 1:
            questionnaires = questionnaires.drop(columns=[col])
            uniform_vars.append(col)
            print(f'Uniform questionnaire item dropped: {col}')

        corr = questionnaires.corr()
        high_corr = []
        for i in range(corr.shape[0]):
            for j in range(i + 1, corr.shape[0]):
                if abs(corr.iloc[i, j]) > 0.85:
                    high_corr.append((corr.index[i], corr.columns[j]))

        # print the highly correlated variables
        if high_corr:
            print('Highly correlated variables:')
            for var in high_corr:
                print(f'{var[0]} and {var[1]} are highly correlated, with a correlation of {corr[var[0]][var[1]]}')

elif choice == "Totals":
    questionnaires = questionnaires[questionnaires_totals['ElementName']]
    col_to_drop = questionnaires[questionnaires.columns[(questionnaires.columns.str.contains('T_SCORE') |
                                                         questionnaires.columns.str.contains('total') |
                                                         questionnaires.columns.str.contains('Total') |
                                                         questionnaires.columns.str.contains('CESD_C_Score') |
                                                         questionnaires.columns.str.contains('_pct'))]].columns
    questionnaires = questionnaires.drop(columns=col_to_drop)
    print(questionnaires.columns)

    # check for correlation between the total scores
    questionnaires = high_corr_checker(questionnaires)
    corr = questionnaires.corr()

    # standardize the data
    print(f'Current data range: {questionnaires.max().max()} - {questionnaires.min().min()}')
    for col in questionnaires.columns:
        # transform the data using min-max scaling
        questionnaires[col] = (questionnaires[col] - questionnaires[col].min()) / (questionnaires[col].max() -
                                                                                   questionnaires[col].min())
    print(f'Standardized data range: {questionnaires.max().max()} - {questionnaires.min().min()}')

else:
    # remove all the total scores
    questionnaires = questionnaires.drop(columns=questionnaires_totals['ElementName'])

    # remove all the text columns
    text_cols = questionnaires_dict[questionnaires_dict['ValueRange'].str.contains('Text', na=False)]['ElementName']
    text_cols = text_cols.str.replace('^CBCL_', '', regex=True)  # to standardize the column names
    questionnaires = questionnaires.drop(columns=text_cols)

    # remove all the PSES columns
    pses_cols = questionnaires[questionnaires.columns[questionnaires.columns.str.contains('PSES')]].columns
    questionnaires = questionnaires.drop(columns=pses_cols)

    # remove all the parent evaluated SCARED columns
    scaredp_cols = questionnaires[questionnaires.columns[questionnaires.columns.str.contains('SCARED_P')]].columns
    scaredc_cols = questionnaires[questionnaires.columns[questionnaires.columns.str.contains('SCARED_C')]].columns

    # calculate the correlation between the parent and child SCARED columns
    scared_corr = []
    for i, j in zip(scaredp_cols, scaredc_cols):
        corr = questionnaires[i].corr(questionnaires[j])
        scared_corr.append((i, j, corr))

    corr_mean = np.mean([corr[2] for corr in scared_corr])  # 0.35866433753391785

    # remove the parent SCARED columns
    questionnaires = questionnaires.drop(columns=scaredp_cols)

    # reconstruct the appetite column
    for index, row in questionnaires.iterrows():
        if row['CDRSR_appetite_type'] == -777:
            questionnaires.at[index, 'CDRSR_appetite'] = 0
        elif row['CDRSR_appetite_type'] == 0:
            questionnaires.at[index, 'CDRSR_appetite'] = 1 * (row['CDRSR_appetite'] - 1)
        elif row['CDRSR_appetite_type'] == 1:
            questionnaires.at[index, 'CDRSR_appetite'] = -1 * (row['CDRSR_appetite'] - 1)

    questionnaires = questionnaires.drop(columns='CDRSR_appetite_type')

    # check nan values
    missing_qn = questionnaires.isna().sum().sort_values(ascending=False)
    missing_qn_var = missing_qn[missing_qn > 0].index
    if missing_qn_var.any():
        for col in missing_qn_var:
            print(f'{col} has {missing_qn[col]} missing values')
    else:
        print('No missing values in the questionnaires data')

    # now, preprocess the data for the bifactor ESEM model
    # remove the id columns
    questionnaires = questionnaires.drop(columns=questionnaires.columns[:5])
    questionnaires = questionnaires.drop(columns=questionnaires.columns[-6:])

    # cbcl_cols = questionnaires.columns[questionnaires.columns.str.contains('Q')]
    # questionnaires = questionnaires[cbcl_cols]

    # if over 99% of the data contains the same value, drop the column
    uniform_vars = []
    for col in questionnaires.columns:
        if len(questionnaires[col].unique()) == 1:
            questionnaires = questionnaires.drop(columns=[col])
            uniform_vars.append(col)
            print(f'Uniform questionnaire item dropped: {col}')

    # check the correlation between the questionnaire items
    questionnaires = high_corr_checker(questionnaires)

    # check again
    high_corr_checker(questionnaires, checker=True)

    # # finally, standardize the data
    # print(f'Current data range: {questionnaires.max().max()} - {questionnaires.min().min()}')
    # for col in questionnaires.columns:
    #     # transform the data using min-max scaling
    #     questionnaires[col] = (questionnaires[col] - questionnaires[col].min()) / (questionnaires[col].max() -
    #                                                                                questionnaires[col].min())
    # print(f'Standardized data range: {questionnaires.max().max()} - {questionnaires.min().min()}')

    # # rename the columns
    # questionnaires.columns = rename_columns(questionnaires.columns, cbcl_cols=False)

    # print the column name according to column index
    for i, col in enumerate(questionnaires.columns):
        print(f'{i + 1}: {col}')

# if __name__ == '__main__':
#     # save cleaned data
#     data.to_csv('./Data/cleaned_data.csv', index=False)
#
#     # save task PLS data
#     psych_without_neuro.to_csv('./Data/PLS_Data/psych_without_neuro.csv', index=False)
#     control_without_neuro.to_csv('./Data/PLS_Data/control_without_neuro.csv', index=False)
#
#     psych_behav.to_csv('./Data/PLS_Data/psych_behav.csv', index=False)
#     control_behav.to_csv('./Data/PLS_Data/control_behav.csv', index=False)
#
#     psych_mean.to_csv('./Data/PLS_Data/psych_mean.csv', index=False)
#     control_mean.to_csv('./Data/PLS_Data/control_mean.csv', index=False)
#
#     psych_median.to_csv('./Data/PLS_Data/psych_median.csv', index=False)
#     control_median.to_csv('./Data/PLS_Data/control_median.csv', index=False)
#
#     psych_combined.to_csv('./Data/PLS_Data/psych_combined.csv', index=False)
#     control_combined.to_csv('./Data/PLS_Data/control_combined.csv', index=False)
#
#     psych_with_mean.to_csv('./Data/PLS_Data/psych_with_mean.csv', index=False)
#     control_with_mean.to_csv('./Data/PLS_Data/control_with_mean.csv', index=False)
#
#     psych_with_median.to_csv('./Data/PLS_Data/psych_with_median.csv', index=False)
#     control_with_median.to_csv('./Data/PLS_Data/control_with_median.csv', index=False)
#
#     psych_with_combined.to_csv('./Data/PLS_Data/psych_with_combined.csv', index=False)
#     control_with_combined.to_csv('./Data/PLS_Data/control_with_combined.csv', index=False)
#
#     # save behavioral PLS data
#     psychopathology.to_csv('./Data/PLS_Data/psychopathology.csv', index=False)
#     behavioral.to_csv('./Data/PLS_Data/behavioral.csv', index=False)
#     neuro_mean_df.to_csv('./Data/PLS_Data/neuro_mean.csv', index=False)
#     neuro_combined_df.to_csv('./Data/PLS_Data/neuro_combined.csv', index=False)
#     iron.to_csv('./Data/PLS_Data/iron.csv', index=False)
# #
#     # save questionnaires data
#     questionnaires.to_csv('./Data/CBCL_customized.csv', index=False)
#     questionnaires.to_csv('./Data/Totals.csv', index=False)


# =============================================================================
#              SGT Data Cleaning (not necessary unless needed)
# =============================================================================
# # SGT
# SGT = SGT.groupby(SGT.index // 2).first()
# # remove unnecessary columns
# SGT = SGT.drop(columns=['Trial', 'QResp'])
#
# list_to_process = ['IDNum', 'Gender', 'Ethnicity', 'Race', 'Age']
# SGT[list_to_process] = SGT[list_to_process].applymap(lambda x: ast.literal_eval(x).get('Q0', None) if x else np.nan)
#
# # explode data
# SGT = SGT.explode(['React', 'Reward', 'keyResponse', 'Bank'])

print('Preprocessing done!')
